from __future__ import annotations

from uuid import UUID
from pathlib import Path

import os
import shutil
from fastapi import FastAPI, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles

from app.ai_prompt import get_ai_system_prompt
from app.master_data import DemoSapMasterDataGateway
from app.models import ApprovalRequest, DocumentStatus, InvoiceDocument, StandardInvoice
from app.ocr import OcrPreview, UploadStore
from app.repository import DatabaseRepository
from app.settings import settings
from app.services import InvoiceService
from app.erp.factory import get_erp_client
from app.routers import admin, master_data, documents

import sys
import json
import logging

log = logging.getLogger(__name__)

if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys._MEIPASS) / "app"
else:
    BASE_DIR = Path(__file__).parent.resolve()

import asyncio
from contextlib import asynccontextmanager

async def periodic_master_data_sync():
    log.info("Starting periodic master data sync loop (every 10 minutes)...")
    while True:
        try:
            await asyncio.sleep(600)  # Wait 10 minutes before first sync and between syncs
            log.info("Executing background ERP master data sync...")
            
            # Run the synchronous function in a background thread to avoid blocking FastAPI
            from app.erp.factory import get_erp_client
            records = await asyncio.to_thread(get_erp_client().sync_all_master_data, None)
            
            if records:
                # Need to update DB
                count = 0
                for r in records:
                    repository.upsert_custom_master_data(
                        category=r["category"],
                        code=r["code"],
                        name=r["name"],
                        extra_data=r["extra_data"],
                        is_default=False
                    )
                    count += 1
                log.info(f"Background sync complete: synchronized {count} master records.")
            
        except asyncio.CancelledError:
            log.info("Periodic master data sync loop stopped.")
            break
        except Exception as e:
            log.error(f"Error during background master data sync: {e}")

@asynccontextmanager
async def lifespan(app: FastAPI):
    sync_task = asyncio.create_task(periodic_master_data_sync())
    yield
    sync_task.cancel()

app = FastAPI(title="AP Invoice OCR", version="0.1.0", description="Validated AP invoice pipeline for SAP Business One drafts.", lifespan=lifespan)
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")

# Include Routers
app.include_router(admin.router)
app.include_router(master_data.router)
app.include_router(documents.router)

repository = DatabaseRepository(
    database_path=settings.data_dir / "invoices.db",
    db_engine=settings.db_engine,
    mysql_host=settings.mysql_host,
    mysql_port=settings.mysql_port,
    mysql_database=settings.mysql_database,
    mysql_user=settings.mysql_user,
    mysql_password=settings.mysql_password,
    mssql_host=settings.mssql_host,
    mssql_port=settings.mssql_port,
    mssql_database=settings.mssql_database,
    mssql_user=settings.mssql_user,
    mssql_password=settings.mssql_password,
)

# Load persistent settings
if (saved_mode := repository.get_setting("posting_mode")):
    settings.posting_mode = saved_mode

upload_store = UploadStore(settings.data_dir, repository=repository)
master_data_gateway = DemoSapMasterDataGateway(repository)
service = InvoiceService(repository, master_data_gateway)


def load_document(document_id: UUID):
    document = repository.get(document_id)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    return document


@app.get("/health")
def health():
    return {
        "status": "ok",
        "sap_service_layer_configured": bool(settings.base_url),
        "sap_draft_posting_enabled": settings.posting_enabled,
    }


@app.get("/ai-prompt")
def ai_prompt():
    return {"system_prompt": get_ai_system_prompt()}


@app.post("/admin/ai-prompt")
def update_ai_prompt(payload: dict):
    new_prompt = payload.get("system_prompt", "")
    if not new_prompt:
        raise HTTPException(status_code=400, detail="system_prompt is required")
    from app.ocr import set_ai_system_prompt
    set_ai_system_prompt(new_prompt)
    return {"status": "ok", "message": "AI model system prompt updated successfully."}


@app.get("/mapping-history")
def mapping_history(limit: int = 50):
    return repository.get_mapping_history(limit=limit)


@app.get("/master-data")
def get_master_data():
    custom = repository.get_custom_master_data()

    custom_by_cat: dict[str, list[dict]] = {}
    for item in custom:
        cat = item["category"]
        if cat not in custom_by_cat:
            custom_by_cat[cat] = []
        custom_by_cat[cat].append(item)

    base_vendors = [{"code": v["code"], "name": v["name"], "extra_data": v.get("extra_data", "")} for v in custom_by_cat.get("vendors", [])]
    
    sap_vendor_details = {v["card_code"]: v for v in repository.get_sap_vendor_master()}
    all_addresses = repository.get_vendor_addresses()
    addr_map = {}
    for addr in all_addresses:
        vc = addr["vendor_code"]
        if vc not in addr_map:
            addr_map[vc] = []
        addr_map[vc].append(addr)
        
    vendors = []
    for bv in base_vendors:
        vc = bv["code"]
        details = sap_vendor_details.get(vc, {})
        bv.update({
            "group_name": details.get("group_name", ""),
            "payment_group": details.get("payment_group", ""),
            "extra_days": details.get("extra_days", 0),
            "currency": details.get("currency", ""),
            "balance": details.get("balance", 0.0),
            "balance_fc": details.get("balance_fc", 0.0),
            "addresses": addr_map.get(vc, [])
        })
        vendors.append(bv)
        
    from app.master_data import DemoSapMasterDataGateway
    import json
    mdg = DemoSapMasterDataGateway(repository)
    existing_codes = {v["code"] for v in vendors}
    for v in mdg.vendors:
        if v.card_code not in existing_codes:
            vendors.append({
                "code": v.card_code,
                "name": v.card_name,
                "extra_data": json.dumps({"gstin": v.gstin}) if v.gstin else "",
                "group_name": "",
                "payment_group": "",
                "extra_days": 0,
                "currency": "INR",
                "balance": 0.0,
                "balance_fc": 0.0,
                "addresses": []
            })
            existing_codes.add(v.card_code)
        
    items = [{"code": i["code"], "name": i["name"], "is_default": i.get("is_default", False)} for i in custom_by_cat.get("items", [])]
    tax_codes = [{"code": t["code"], "rate": t["extra_data"] or "0", "is_default": t.get("is_default", False)} for t in custom_by_cat.get("tax_codes", [])]
    cc1 = [{"code": c["code"], "name": c["name"], "is_default": c.get("is_default", False)} for c in custom_by_cat.get("cost_centers1", [])]
    cc2 = [{"code": c["code"], "name": c["name"], "is_default": c.get("is_default", False)} for c in custom_by_cat.get("cost_centers2", [])]
    cc3 = [{"code": c["code"], "name": c["name"], "is_default": c.get("is_default", False)} for c in custom_by_cat.get("cost_centers3", [])]
    accounts = [{"code": a["code"], "name": a["name"], "is_default": a.get("is_default", False)} for a in custom_by_cat.get("accounts", [])]

    currencies = [{"code": c["code"], "name": c["name"], "is_default": c.get("is_default", False)} for c in custom_by_cat.get("currencies", [])]
    branches = [{"code": b["code"], "name": b["name"], "is_default": b.get("is_default", False)} for b in custom_by_cat.get("branches", [])]
    series = [{"code": s["code"], "name": s["name"], "extra_data": s.get("extra_data") or "", "is_default": s.get("is_default", False)} for s in custom_by_cat.get("series", [])]

    # Fix: Fetch WTax Codes and Vendor Addresses from the custom master data (synced from SAP)
    wtax_codes = custom_by_cat.get("wtax_codes", [])
    if not wtax_codes:
        wtax_codes = [{"code": c["wtax_code"], "name": c["description"], "extra_data": c["wtax_rate"]} for c in repository.get_wtax_codes(active_only=True)]
        
    vendor_addresses = custom_by_cat.get("vendor_addresses", [])
    if not vendor_addresses:
        vendor_addresses = [{"code": f"{a['vendor_code']}::{a['address_code']}", "name": a["vendor_code"], "extra_data": '{"address_type":"' + a['address_type'] + '","address_text":"' + a['address_text'] + '","block":"' + a['block'] + '","building":"' + a['building'] + '","street":"' + a['street'] + '","street_no":"' + a['street_no'] + '","city":"' + a['city'] + '","state":"' + a['state'] + '","country":"' + a['country'] + '","gst_regn_no":"' + a['gst_regn_no'] + '","is_default":' + str(a['is_default']).lower() + '}'} for a in repository.get_vendor_addresses()]
    sac_entries = custom_by_cat.get("sac_entries", [])

    return {
        "tax_codes": tax_codes,
        "cost_centers1": cc1,
        "cost_centers2": cc2,
        "cost_centers3": cc3,
        "vendors": vendors,
        "vendor_addresses": vendor_addresses,
        "items": items,
        "currencies": currencies,
        "branches": branches,
        "series": series,
        "accounts": accounts,
        "wtax_codes": wtax_codes,
        "sac_entries": sac_entries,
        "place_of_supply": custom_by_cat.get("place_of_supply", []),
        "locations": custom_by_cat.get("locations", []),
    }


@app.get("/admin/ai-train/stats")
def get_ai_training_stats():
    return repository.get_ai_training_stats()


@app.post("/admin/ai-train/upload-dump")
async def upload_dump_training_dataset(file: UploadFile = File(...)):
    try:
        content = await file.read()
        return service.train_ai_from_file_dump(content, file.filename or "dump.csv")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to process dump file: {exc}")


@app.post("/admin/ai-train/sync-sap-history")
def sync_sap_history_training(top: int = 100):
    try:
        return service.train_ai_from_sap_history(top=top)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to sync SAP history: {exc}")


@app.post("/auth/login")
def login(payload: dict):
    username = payload.get("username", "")
    password = payload.get("password", "")
    role = payload.get("role", "")

    user_info = repository.verify_and_get_user(username, password, role)
    if user_info:
        return {"status": "ok", "username": user_info["username"], "role": user_info["role"]}
    raise HTTPException(status_code=401, detail="Invalid username or password. Please try 'admin' / 'admin' or 'user' / 'user'.")





@app.get("/admin/validation-rules")
def get_validation_rules():
    return repository.get_validation_rules()

@app.post("/admin/validation-rules")
def upsert_validation_rule(payload: dict):
    rule_id = payload.get("id")
    rule_name = payload.get("rule_name")
    target_field = payload.get("target_field")
    condition = payload.get("condition")
    condition_value = payload.get("condition_value", "")
    error_message = payload.get("error_message")
    is_active = payload.get("is_active", True)
    
    if not all([rule_name, target_field, condition, error_message]):
        raise HTTPException(status_code=400, detail="Missing required validation rule fields")
        
    return repository.upsert_validation_rule(
        rule_id, rule_name, target_field, condition, condition_value, error_message, is_active
    )

@app.delete("/admin/validation-rules/{rule_id}")
def delete_validation_rule(rule_id: int):
    repository.delete_validation_rule(rule_id)
    return {"status": "ok"}

@app.get("/admin/config")
def get_admin_config():
    return {
        "tesseract_cmd": settings.tesseract_cmd,
        "posting_enabled": settings.posting_enabled,
        "posting_mode": settings.posting_mode,
        "ocr_provider": settings.ocr_provider,
        "ocr_api_key": settings.ocr_api_key,
        "ocr_api_url": settings.ocr_api_url,
        "ocr_model_name": settings.ocr_model_name,
    }


@app.post("/admin/config")
def update_admin_config(payload: dict):
    if "posting_enabled" in payload:
        settings.posting_enabled = bool(payload["posting_enabled"])
    if "posting_mode" in payload and payload["posting_mode"]:
        settings.posting_mode = payload["posting_mode"].lower()
        repository.set_setting("posting_mode", settings.posting_mode)

    if "ocr_provider" in payload and payload["ocr_provider"]:
        settings.ocr_provider = payload["ocr_provider"]
        os.environ["OCR_PROVIDER"] = payload["ocr_provider"]
    if "ocr_api_key" in payload:
        settings.ocr_api_key = payload["ocr_api_key"]
        os.environ["NVIDIA_API_KEY"] = payload["ocr_api_key"]
        os.environ["OCR_API_KEY"] = payload["ocr_api_key"]
    if "ocr_api_url" in payload:
        settings.ocr_api_url = payload["ocr_api_url"]
        os.environ["NVIDIA_OCR_URL"] = payload["ocr_api_url"]
        os.environ["OCR_API_URL"] = payload["ocr_api_url"]
    if "ocr_model_name" in payload and payload["ocr_model_name"]:
        settings.ocr_model_name = payload["ocr_model_name"]
        os.environ["OCR_MODEL_NAME"] = payload["ocr_model_name"]

    return {"status": "ok", "message": "System & OCR Configuration updated successfully."}


@app.post("/api/open-documents/sync")
def sync_all_open_documents(doc_type: str = Query("PO")):
    """Sync all Open POs or GRNs from SAP to local DB based on doc_type."""
    try:
        if doc_type not in ["PO", "GRN"]:
            return {"status": "error", "detail": "Invalid doc_type. Must be PO or GRN."}
            
        sap_data = get_erp_client().get_all_open_documents(doc_type)
        
        doc_list = []
        for doc in sap_data:
            doc_list.append({
                "doc_entry": doc.get("DocEntry"),
                "doc_num": str(doc.get("DocNum")),
                "vendor_code": doc.get("CardCode"),
                "doc_date": doc.get("DocDate"),
                "total_amount": float(doc.get("DocTotal", 0.0) or 0.0),
                "lines_payload": json.dumps(doc)
            })
            
        repository.replace_open_documents(doc_type, doc_list)
        return {"status": "ok", "message": f"Synced {len(doc_list)} {doc_type}s from SAP."}
    except Exception as e:
        log.error(f"Error syncing {doc_type} documents: {e}")
        return {"status": "error", "detail": str(e)}


@app.get("/api/open-documents")
def get_all_open_documents():
    """Fetch all Open POs and GRNs across all vendors."""
    docs = repository.get_all_open_documents()
    return {"status": "ok", "documents": docs}


@app.get("/api/vendors/{vendor_code}/open-documents")
def get_vendor_open_documents(vendor_code: str, doc_type: str = Query(..., description="PO or GRN")):
    """Fetch Open POs or Open GRNs for a given vendor code."""
    doc_type = doc_type.upper()
    if doc_type == "PO":
        docs = repository.get_open_pos(vendor_code)
    elif doc_type == "GRN":
        docs = repository.get_open_grns(vendor_code)
    else:
        raise HTTPException(status_code=400, detail="Invalid doc_type. Must be PO or GRN.")
    
    return {"status": "ok", "documents": docs}



@app.get("/api/branding")
def get_branding():
    app_name = repository.get_setting("app_name") or "SAP Business One OCR Extractor"
    logo_url = repository.get_setting("logo_url") or "/static/client_logo.png"
    return {"app_name": app_name, "logo_url": logo_url}


@app.post("/admin/branding")
def update_branding(app_name: str = Form(None), logo: UploadFile = File(None)):
    if app_name:
        repository.set_setting("app_name", app_name)
    
    if logo and logo.filename:
        ext = os.path.splitext(logo.filename)[1]
        if not ext:
            ext = ".png"
        filename = f"client_logo{ext}"
        filepath = str(BASE_DIR / "static" / filename)
        with open(filepath, "wb") as f:
            shutil.copyfileobj(logo.file, f)
        
        repository.set_setting("logo_url", f"/static/{filename}")
    
    return {"status": "ok", "message": "Branding updated successfully."}


@app.get("/admin/form-fields")
def get_form_fields():
    return repository.get_form_fields()


@app.post("/admin/form-fields")
def save_form_field(payload: dict):
    field_id = payload.get("field_id")
    section = payload.get("section", "header")
    sap_param_name = payload.get("sap_param_name")
    label = payload.get("label")
    field_type = payload.get("field_type", "text")
    enabled = payload.get("enabled", True)
    required = payload.get("required", False)
    sort_order = payload.get("sort_order", 0)
    visible = payload.get("visible", True)

    if not field_id or not sap_param_name or not label:
        raise HTTPException(status_code=400, detail="Field ID, SAP Parameter Name, and Label are required.")

    repository.upsert_form_field(field_id, section, sap_param_name, label, field_type, enabled, required, sort_order, visible)
    return {"status": "ok", "message": f"Saved field '{label}' ({sap_param_name})"}


@app.delete("/admin/form-fields")
def delete_form_field(field_id: str):
    repository.delete_form_field(field_id)
    return {"status": "ok", "message": f"Deleted field '{field_id}'"}





@app.post("/admin/master-data/bulk-upload")
async def bulk_upload_master_data_excel(file: UploadFile = File(...)):
    import io
    import openpyxl

    if not file.filename.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="Only Excel (.xlsx/.xls) or CSV files are supported.")

    contents = await file.read()
    records_added = 0

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                continue

            header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]

            sheet_category = None
            s_clean = sheet_name.strip().lower()
            if "vendor" in s_clean or "bp" in s_clean or "supplier" in s_clean:
                sheet_category = "vendors"
            elif "item" in s_clean or "product" in s_clean:
                sheet_category = "items"
            elif "tax" in s_clean:
                sheet_category = "tax_codes"
            elif "cost" in s_clean or "center" in s_clean:
                sheet_category = "cost_centers1"
            elif "account" in s_clean or "gl" in s_clean:
                sheet_category = "accounts"
            elif "branch" in s_clean:
                sheet_category = "branches"

            for row in rows[1:]:
                if not any(row):
                    continue
                row_dict = {header[i]: str(row[i]).strip() if i < len(row) and row[i] is not None else "" for i in range(len(header))}
                
                cat = row_dict.get("category") or sheet_category or "custom"
                code = (
                    row_dict.get("code") or 
                    row_dict.get("cardcode") or 
                    row_dict.get("itemcode") or 
                    row_dict.get("accountcode") or 
                    row_dict.get("account_code") or 
                    row_dict.get("gl_account") or 
                    row_dict.get("acctcode") or 
                    row_dict.get("taxcode") or 
                    row_dict.get("costingcode")
                )
                name = (
                    row_dict.get("name") or 
                    row_dict.get("cardname") or 
                    row_dict.get("itemname") or 
                    row_dict.get("accountname") or 
                    row_dict.get("account_name") or 
                    row_dict.get("acctname") or 
                    row_dict.get("description") or 
                    code or 
                    ""
                )
                extra = row_dict.get("extra_data") or row_dict.get("rate") or row_dict.get("gstin") or row_dict.get("formatcode") or ""

                if cat and code:
                    repository.upsert_custom_master_data(cat, code, name, extra, False)
                    records_added += 1

        return {"status": "ok", "message": f"Successfully imported {records_added} master data records from Excel workbook!", "count": records_added}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to process Excel file: {exc}")


@app.get("/admin/master-data/template")
def download_master_data_template():
    import io
    import openpyxl
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "MasterData"
    ws1.append(["Category", "Code", "Name", "Extra_Data"])
    ws1.append(["vendors", "V01145", "Super Tech Solutions Pvt Ltd", "33ABCDE1234F1Z5"])
    ws1.append(["items", "RM0099", "Industrial Steel Grade A", ""])
    ws1.append(["tax_codes", "GST18_NEW", "GST 18% Custom", "18.00"])
    ws1.append(["accounts", "5217999", "Custom Operational Expense Account", ""])
    ws1.append(["cost_centers1", "TN-CHE", "Chennai Production Center", ""])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {"Content-Disposition": "attachment; filename=Master_Data_Import_Template.xlsx"}
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)



import pathlib
import pathlib

@app.get("/", include_in_schema=False)
def dashboard():
    return FileResponse(BASE_DIR / "static" / "index.html")


@app.get("/manual", include_in_schema=False)
def user_manual():
    return FileResponse(BASE_DIR / "static" / "manual.html")



@app.post("/uploads/ocr", response_model=OcrPreview, status_code=status.HTTP_201_CREATED)
async def upload_invoice(file: UploadFile = File(...)):
    """Persist source file then return OCR-derived data for human preview and correction."""
    return await upload_store.save_and_extract(file)


@app.post("/uploads/ocr/bulk", status_code=status.HTTP_201_CREATED)
async def bulk_upload_invoices(files: list[UploadFile] = File(...)):
    """Process multiple invoice files at once, save extracted documents to Queue, and return status."""
    results = []
    for file in files:
        try:
            preview = await upload_store.save_and_extract(file)
            if preview.invoice:
                doc = InvoiceDocument(invoice=preview.invoice, source_filename=preview.filename)
                doc = service.map_document(doc)
                doc = service.validate(doc)
                repository.save(doc)
                results.append({
                    "filename": file.filename,
                    "status": "SUCCESS",
                    "document_id": str(doc.document_id),
                    "invoice_number": doc.invoice.invoice_header.invoice_number,
                    "supplier_name": doc.invoice.invoice_header.supplier_name,
                    "grand_total": float(doc.invoice.totals.grand_total),
                    "doc_status": str(doc.status),
                })
            else:
                results.append({
                    "filename": file.filename,
                    "status": "FAILED",
                    "message": "Failed to extract invoice data",
                })
        except Exception as exc:
            results.append({
                "filename": file.filename,
                "status": "FAILED",
                "message": str(exc),
            })
    return {"total": len(files), "processed": len(results), "results": results}


@app.post("/uploads/{upload_id}/correction", response_model=OcrPreview)
def save_upload_correction(upload_id: UUID, invoice: StandardInvoice):
    return upload_store.save_correction(upload_id, invoice)


@app.get("/uploads/{upload_id}", response_model=OcrPreview)
def get_upload_preview(upload_id: UUID):
    upload = upload_store.get_candidate(upload_id)
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    invoice = None
    if upload["invoice_payload"]:
        invoice = StandardInvoice.model_validate(upload["invoice_payload"])
    return OcrPreview(
        upload_id=upload_id,
        filename=upload["filename"],
        extraction_status=upload["extraction_status"],
        message="Upload preview retrieved.",
        extracted_text=upload["extracted_text"],
        invoice=invoice,
        field_confidence=invoice.field_confidence if invoice else {},
    )


@app.get("/uploads/{upload_id}/file")
def get_upload_file(upload_id: UUID):
    upload = upload_store.get_candidate(upload_id)
    if not upload or not upload.get("stored_path"):
        raise HTTPException(status_code=404, detail="Upload file not found")
    file_path = Path(upload["stored_path"])
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Physical upload file missing")
    return FileResponse(file_path, media_type=upload.get("content_type") or "application/pdf")


@app.post("/documents", response_model=InvoiceDocument, status_code=status.HTTP_201_CREATED)
def create_document(invoice: StandardInvoice, source_filename: str | None = None):
    doc = InvoiceDocument(invoice=invoice, source_filename=source_filename)
    doc = service.map_document(doc)
    doc = service.validate(doc)
    return repository.save(doc)






